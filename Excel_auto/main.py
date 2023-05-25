import sys
import pandas as pd
from openpyxl.reader.excel import load_workbook

input = sys.stdin.readline


def all_value_check(all_values):
    for r in all_values:
        print(r)


def value_check_and_expert(standard, all_values, standard_col, standard_value):
    for v in zip(standard_col, standard_value):
        print(v)

    res = list()
    res.append(standard)
    for v in zip(standard_col, standard_value):
        for r in all_values:
            if r[int(v[0])] == v[1]:
                res.append(r)

    for r in res:
        print(r)

    df = pd.DataFrame.from_records(res)
    print("추출할 경로를 입력하세요.")
    export_path = input().strip() + "/test.xlsx"
    df.to_excel(export_path)


def main():
    print("파일 경로을 입력해주세요.")
    excel_path = input().strip() + ".xlsx"
    print("시트 이름을 입력해주세요.")
    sheet_name = input().strip()

    load_wb = load_workbook(excel_path, data_only=True)
    load_ws = load_wb[sheet_name]

    column_max = load_ws.max_column
    row_max = load_ws.max_row

    all_values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)

    standard = all_values[0]
    print(standard)

    # 엑셀 내 모든 값 확인
    # all_value_check(all_values)

    print("Column을 입력해주세요. [복수 가능]")
    standard_col = input().split()

    print("Column에 대응한 값을 입력해주세요. [Column의 수와 동일]")
    standard_value = input().split()

    value_check_and_expert(standard, all_values, standard_col, standard_value)


main()


